# OTP Relay Server - main.py
# Stack: FastAPI + Python 3.12 + Exchange SMTP (internal only)
# No external APIs. Runs entirely on your company LAN.
#
# Delivery model: OTP is displayed on-screen via polling. Email is NOT used
# for OTP delivery. SMTP config and /admin/smtp-test are retained for
# diagnostics only.

import asyncio
import json
import logging
import os
import re
import secrets
import smtplib
from collections import deque
from datetime import datetime, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from io import BytesIO
from typing import Any, Dict, List, Optional

import bcrypt
import openpyxl
from dotenv import load_dotenv
from fastapi import FastAPI, File, Header, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

BASE_DIR = Path(__file__).resolve().parent
FRONTEND_DIR = BASE_DIR / "frontend"


def _resolve_runtime_path(value: str) -> Path:
    path = Path(value)
    return path if path.is_absolute() else BASE_DIR / path


load_dotenv(BASE_DIR / ".env")

app = FastAPI(title="OTP Relay")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # safe for current LAN-only deployment
    allow_methods=["*"],
    allow_headers=["*"],
)

# -- Config -------------------------------------------------------------------
SMS_SECRET_TOKEN = os.getenv("SMS_SECRET_TOKEN", "changeme")

SMTP_HOST = os.getenv("SMTP_HOST", "mail.company.local")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "otp-relay@company.com")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "")
SMTP_USE_TLS = os.getenv("SMTP_USE_TLS", "true").lower() == "true"
SMTP_AUTH = os.getenv("SMTP_AUTH", "true").lower() == "true"
FROM_EMAIL = os.getenv("FROM_EMAIL", SMTP_USER)
FROM_NAME = os.getenv("FROM_NAME", "OTP Relay")

# How long the active user has to trigger their OTP before being evicted.
# Other users wait until this window expires or OTP is delivered.
CLAIM_EXPIRY_SEC = int(os.getenv("CLAIM_EXPIRY_SEC", "90"))

# How long the delivered OTP stays visible on-screen before being purged.
OTP_DISPLAY_SEC = int(os.getenv("OTP_DISPLAY_SEC", "285"))  # 4 min 45 sec

# If two claims arrive within this window, log a concurrent_risk event.
CONCURRENT_RISK_SEC = int(os.getenv("CONCURRENT_RISK_SEC", "30"))

USERS_EXCEL_PATH = str(_resolve_runtime_path(os.getenv("USERS_EXCEL_PATH", "data/users.xlsx")))
USERS_EXCEL_MAX_BYTES = int(os.getenv("USERS_EXCEL_MAX_BYTES", str(5 * 1024 * 1024)))
AUDIT_LOG_PATH = str(_resolve_runtime_path(os.getenv("AUDIT_LOG_PATH", "data/audit.log")))

# -- State --------------------------------------------------------------------
users: Dict[str, Dict[str, str]] = {}
claim_queue: deque = deque()

# Delivered OTPs held in memory only - never written to disk or logs.
# Structure: { token: { "otp": str, "arrived_at": datetime } }
pending_otps: Dict[str, Dict[str, Any]] = {}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%dT%H:%M:%SZ",
)
logger = logging.getLogger("otp-relay")

# -- Server-backed wizard/admin state -----------------------------------------
DATA_DIR = _resolve_runtime_path(os.environ.get("OTP_RELAY_DATA_DIR", "data"))
WIZARD_FILE = DATA_DIR / "wizard_progress.json"
AUTH_FILE = DATA_DIR / "admin_auth.json"
CONFIG_FILE = DATA_DIR / "admin_config.json"
DEFAULT_ADMIN_TOKENS = ["JPR", "AMD", "SCH"]
ADMIN_TTL_SECONDS = 8 * 60 * 60
ADMIN_SESSIONS: Dict[str, float] = {}


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _ensure_data_dir() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)


def _read_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        raw = path.read_text(encoding="utf-8").strip()
        if not raw:
            return default
        return json.loads(raw)
    except Exception as exc:
        logger.warning("Could not read %s: %s", path, exc)
        return default


def _write_json(path: Path, payload: Any) -> None:
    """Write JSON atomically so a pod restart cannot leave a half-written file."""
    _ensure_data_dir()
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    tmp_path.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    tmp_path.replace(path)


def _wizard_db() -> Dict[str, dict]:
    return _read_json(WIZARD_FILE, {})


def _save_wizard_db(db: Dict[str, dict]) -> None:
    _write_json(WIZARD_FILE, db)


def _auth_db() -> Dict[str, Any]:
    return _read_json(AUTH_FILE, {})


def _save_auth_db(db: Dict[str, Any]) -> None:
    _write_json(AUTH_FILE, db)


def _config_db() -> Dict[str, Any]:
    env_tokens = os.environ.get("ADMIN_TOKENS", "")
    env_default = [t.strip().upper() for t in env_tokens.split(",") if t.strip()] or DEFAULT_ADMIN_TOKENS
    return _read_json(CONFIG_FILE, {"admin_tokens": env_default})


def _save_config_db(db: Dict[str, Any]) -> None:
    _write_json(CONFIG_FILE, db)


def _purge_admin_sessions() -> None:
    now_ts = datetime.now(timezone.utc).timestamp()
    stale = [session for session, ts in ADMIN_SESSIONS.items() if now_ts - ts > ADMIN_TTL_SECONDS]
    for session in stale:
        ADMIN_SESSIONS.pop(session, None)


def _require_admin(session: Optional[str]) -> None:
    _purge_admin_sessions()
    if not session:
        raise HTTPException(status_code=401, detail="Missing admin session")
    ts = ADMIN_SESSIONS.get(session)
    if not ts:
        raise HTTPException(status_code=401, detail="Invalid admin session")
    ADMIN_SESSIONS[session] = datetime.now(timezone.utc).timestamp()


def _model_dump(model: BaseModel) -> Dict[str, Any]:
    """Support both Pydantic v1 and v2."""
    if hasattr(model, "model_dump"):
        return model.model_dump()
    return model.dict()


class WizardRecord(BaseModel):
    token: str
    display_name: str = ""
    iits_username: str = ""
    adm_username: str = ""
    completed: List[str] = Field(default_factory=list)
    adminCompleted: List[str] = Field(default_factory=list)
    iits_pw_date: Optional[str] = None
    adm_pw_date: Optional[str] = None
    vpn_date: Optional[str] = None


class CredentialPayload(BaseModel):
    credential: str
    current: Optional[str] = None


class ConfigPayload(BaseModel):
    admin_tokens: List[str]


# -- User loading --------------------------------------------------------------
def load_users_from_excel(path: str, replace_existing: bool = True) -> int:
    """
    Reads users.xlsx. Expected columns (row 1 = headers):
      token - 2 or 3 character unique string, e.g. AH or AHM
      name  - display name
      email - company email address
    Column names are case-insensitive.
    Skipped rows are written to the audit log so IT can fix them.
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    raw_headers = [
        str(cell.value).strip().lower() if cell.value else ""
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]

    required_headers = {"token", "name", "email"}
    missing_headers = sorted(required_headers - set(raw_headers))
    if missing_headers:
        raise ValueError(f"users.xlsx missing required column(s): {', '.join(missing_headers)}")

    loaded = 0
    skipped = 0
    seen_tokens: Dict[str, int] = {}
    imported_users: Dict[str, Dict[str, str]] = {}

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(value is None for value in row):
            continue

        row_dict = dict(zip(raw_headers, row))
        token = str(row_dict.get("token", "") or "").strip().upper()
        name = str(row_dict.get("name", "") or "").strip()
        email = str(row_dict.get("email", "") or "").strip()

        if len(token) == 0:
            audit("import_skipped", detail=f"Row {row_num}: empty token - name={repr(name)} email={repr(email)}", status="warn")
            skipped += 1
            continue

        if not (2 <= len(token) <= 3):
            audit("import_skipped", token=token, detail=f"Row {row_num}: token must be 2 or 3 characters, got {len(token)} ({repr(token)})", status="warn")
            skipped += 1
            continue

        if not re.match(r"^[A-Z0-9]+$", token):
            audit("import_skipped", token=token, detail=f"Row {row_num}: token contains invalid characters ({repr(token)}) - only letters and digits allowed", status="warn")
            skipped += 1
            continue

        if not email:
            audit("import_skipped", token=token, detail=f"Row {row_num}: missing email address for {repr(name)}", status="warn")
            skipped += 1
            continue

        if "@" not in email:
            audit("import_skipped", token=token, detail=f"Row {row_num}: invalid email address {repr(email)}", status="warn")
            skipped += 1
            continue

        if token in seen_tokens:
            audit("import_skipped", token=token, detail=f"Row {row_num}: duplicate token - already defined at row {seen_tokens[token]}", status="warn")
            skipped += 1
            continue

        seen_tokens[token] = row_num
        imported_users[token] = {"token": token, "name": name, "email": email}
        loaded += 1

    if replace_existing:
        users.clear()
        users.update(imported_users)

    logger.info("Loaded %s users from %s (%s rows skipped)", loaded, path, skipped)
    if skipped > 0:
        audit("import_complete", detail=f"{loaded} users loaded, {skipped} rows skipped - check import_skipped entries above", status="warn")
    else:
        audit("import_complete", detail=f"{loaded} users loaded, no issues")
    return loaded


# -- Audit log ----------------------------------------------------------------
def audit(event: str, token: Optional[str] = None, detail: str = "", status: str = "info") -> None:
    entry = {
        "ts": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "event": event,
        "token": token or "",
        "detail": detail,
        "status": status,
    }
    try:
        audit_path = Path(AUDIT_LOG_PATH)
        audit_path.parent.mkdir(parents=True, exist_ok=True)
        with audit_path.open("a", encoding="utf-8") as handle:
            handle.write(json.dumps(entry) + "\n")
    except Exception as exc:
        logger.warning("Could not write audit log: %s", exc)

    level = {"info": logging.INFO, "warn": logging.WARNING, "error": logging.ERROR}.get(status, logging.INFO)
    logger.log(level, "[%s] token=%s  %s", event, token or "-", detail)


def read_audit_log(limit: int = 200) -> list:
    try:
        lines = Path(AUDIT_LOG_PATH).read_text(encoding="utf-8").strip().splitlines()
        entries = [json.loads(line) for line in lines if line.strip()]
        return list(reversed(entries))[:limit]
    except FileNotFoundError:
        return []
    except Exception as exc:
        logger.warning("Could not read audit log: %s", exc)
        return []


# -- Queue and OTP state helpers ----------------------------------------------
def purge_expired() -> None:
    """Evict the front-of-queue claim if it has exceeded CLAIM_EXPIRY_SEC."""
    now = datetime.utcnow()
    while claim_queue:
        age = (now - claim_queue[0]["claimed_at"]).total_seconds()
        if age > CLAIM_EXPIRY_SEC:
            expired = claim_queue.popleft()
            audit("claim_expired", expired["token"], f"No OTP arrived within {CLAIM_EXPIRY_SEC}s - evicted from slot 1", "warn")
        else:
            break


def purge_stale_otps() -> None:
    """Remove delivered OTPs that have exceeded OTP_DISPLAY_SEC."""
    now = datetime.utcnow()
    stale = [
        token for token, value in pending_otps.items()
        if (now - value["arrived_at"]).total_seconds() > OTP_DISPLAY_SEC
    ]
    for token in stale:
        del pending_otps[token]
        audit("otp_display_expired", token, f"OTP display window closed after {OTP_DISPLAY_SEC}s")


def extract_otp(text: str) -> str:
    match = re.search(r"\b\d{4,8}\b", text)
    return match.group() if match else "-"


# -- Email diagnostics ---------------------------------------------------------
def send_email(to_email: str, name: str, subject: str, html: str) -> None:
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = f"{FROM_NAME} <{FROM_EMAIL}>"
    msg["To"] = to_email
    msg.attach(MIMEText(html, "html"))

    if SMTP_USE_TLS:
        server = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=20)
        server.ehlo()
        server.starttls()
    else:
        server = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=20)

    try:
        if SMTP_AUTH:
            server.login(SMTP_USER, SMTP_PASSWORD)
        server.sendmail(FROM_EMAIL, to_email, msg.as_string())
    finally:
        server.quit()


# -- Background task -----------------------------------------------------------
async def background_purge() -> None:
    """Runs every 15 seconds to expire stale queue entries and OTP display windows."""
    while True:
        await asyncio.sleep(15)
        purge_expired()
        purge_stale_otps()


# -- Endpoints ----------------------------------------------------------------
@app.on_event("startup")
async def startup() -> None:
    _ensure_data_dir()
    if os.path.exists(USERS_EXCEL_PATH):
        count = load_users_from_excel(USERS_EXCEL_PATH)
        audit("server_start", detail=f"{count} users loaded")
    else:
        logger.warning("users.xlsx not found at %s", USERS_EXCEL_PATH)
        audit("server_start", detail="No users.xlsx - POST /admin/reload-users after adding it", status="warn")
    asyncio.create_task(background_purge())


@app.get("/healthz")
async def healthz():
    return {"status": "ok"}


@app.get("/readyz")
async def readyz():
    return {"status": "ok", "users_loaded": len(users)}


@app.post("/claim-otp")
async def claim_otp(request: Request):
    data = await request.json()
    token = str(data.get("token", "")).strip().upper()

    if token not in users:
        audit("claim_rejected", token, "Unknown token", "error")
        raise HTTPException(status_code=404, detail="Token not recognised. Check with your IT department.")

    purge_expired()
    purge_stale_otps()

    for i, claim in enumerate(claim_queue):
        if claim["token"] == token:
            age = (datetime.utcnow() - claim["claimed_at"]).total_seconds()
            remaining = max(0, int(CLAIM_EXPIRY_SEC - age))
            audit("claim_duplicate", token, f"Already at position {i + 1}", "warn")
            return {
                "status": "already_queued",
                "position": i + 1,
                "expires_in": remaining,
                "queue_depth": len(claim_queue),
            }

    if token in pending_otps:
        age = (datetime.utcnow() - pending_otps[token]["arrived_at"]).total_seconds()
        remaining = max(0, int(OTP_DISPLAY_SEC - age))
        return {"status": "otp_ready", "expires_in": remaining}

    now = datetime.utcnow()

    if claim_queue:
        front_age = (now - claim_queue[0]["claimed_at"]).total_seconds()
        if front_age < CONCURRENT_RISK_SEC:
            audit(
                "concurrent_risk",
                token,
                f"New claim while {claim_queue[0]['token']} has been active for only {int(front_age)}s",
                "warn",
            )

    claim_queue.append({
        "token": token,
        "name": users[token]["name"],
        "email": users[token]["email"],
        "claimed_at": now,
    })

    position = len(claim_queue)
    queue_depth = len(claim_queue)
    wait_estimate = max(0, (position - 1) * CLAIM_EXPIRY_SEC)

    audit("claim_queued", token, f"Queue position {position} of {queue_depth}")
    return {
        "status": "queued",
        "position": position,
        "name": users[token]["name"],
        "expires_in": CLAIM_EXPIRY_SEC,
        "queue_depth": queue_depth,
        "wait_estimate": wait_estimate,
    }


@app.get("/claim-status/{token}")
async def claim_status(token: str):
    token = token.upper()

    purge_expired()
    purge_stale_otps()

    if token in pending_otps:
        age = (datetime.utcnow() - pending_otps[token]["arrived_at"]).total_seconds()
        remaining = max(0, int(OTP_DISPLAY_SEC - age))
        return {"status": "delivered", "otp": pending_otps[token]["otp"], "expires_in": remaining}

    for i, claim in enumerate(claim_queue):
        if claim["token"] == token:
            age = (datetime.utcnow() - claim["claimed_at"]).total_seconds()
            remaining = max(0, int(CLAIM_EXPIRY_SEC - age))
            wait_estimate = max(0, i * CLAIM_EXPIRY_SEC)
            return {
                "status": "waiting",
                "position": i + 1,
                "expires_in": remaining,
                "queue_depth": len(claim_queue),
                "wait_estimate": wait_estimate,
            }

    for entry in read_audit_log(500):
        if entry.get("token") == token:
            if entry["event"] in ("otp_delivered", "otp_display_expired"):
                return {"status": "done"}
            if entry["event"] == "claim_expired":
                return {"status": "idle_expired"}
            break

    return {"status": "unknown"}


@app.delete("/claim-otp/{token}")
async def cancel_claim(token: str):
    """Discard a delivered OTP and remove/requeue claim state for retry flows."""
    token = token.upper()

    if token in pending_otps:
        del pending_otps[token]
        audit("otp_discarded", token, "User requested retry - OTP discarded from memory")

    global claim_queue
    before = len(claim_queue)
    claim_queue = deque(claim for claim in claim_queue if claim["token"] != token)
    if len(claim_queue) < before:
        audit("claim_cancelled", token, "Removed from queue by user")

    return {"status": "ok"}


@app.post("/sms-received")
async def sms_received(request: Request):
    if request.headers.get("X-Secret-Token", "") != SMS_SECRET_TOKEN:
        audit("sms_rejected", detail="Wrong secret token", status="error")
        raise HTTPException(status_code=401)

    data = await request.json()
    sms_body = str(data.get("body", "")).strip()
    audit("sms_received", detail=f"SMS arrived ({len(sms_body)} chars)")

    purge_expired()
    purge_stale_otps()

    if not claim_queue:
        await asyncio.sleep(4)
        purge_expired()
        if not claim_queue:
            audit("sms_unmatched", detail="No claimant in queue - SMS discarded", status="warn")
            return {"status": "no_claimant"}

    recipient = claim_queue.popleft()
    otp = extract_otp(sms_body)

    pending_otps[recipient["token"]] = {"otp": otp, "arrived_at": datetime.utcnow()}

    audit("otp_delivered", recipient["token"], "OTP ready for display - queue unblocked")
    return {"status": "delivered", "recipient": recipient["name"]}


@app.get("/admin/log")
async def get_log(limit: int = 200):
    entries = read_audit_log(limit)
    return {"entries": entries, "total": len(entries)}


@app.get("/admin/queue")
async def get_queue():
    now = datetime.utcnow()
    return {
        "queue": [
            {
                "token": claim["token"],
                "name": claim["name"],
                "email": claim["email"],
                "claimed_at": claim["claimed_at"].strftime("%Y-%m-%dT%H:%M:%SZ"),
                "expires_in": max(0, int(CLAIM_EXPIRY_SEC - (now - claim["claimed_at"]).total_seconds())),
                "position": i + 1,
            }
            for i, claim in enumerate(claim_queue)
        ]
    }


@app.get("/admin/users")
async def list_users():
    return {
        "count": len(users),
        "users": [{"token": user["token"], "name": user["name"], "email": user["email"]} for user in users.values()],
    }




@app.get("/admin/users/status")
async def users_file_status(x_admin_session: Optional[str] = Header(default=None)):
    _require_admin(x_admin_session)
    path = Path(USERS_EXCEL_PATH)
    exists = path.exists()
    stat = path.stat() if exists else None
    return {
        "exists": exists,
        "path": str(path),
        "users_loaded": len(users),
        "size_bytes": stat.st_size if stat else 0,
        "updated_at": datetime.fromtimestamp(stat.st_mtime, timezone.utc).isoformat() if stat else None,
        "max_size_bytes": USERS_EXCEL_MAX_BYTES,
    }


@app.post("/admin/users/upload")
async def upload_users_excel(file: UploadFile = File(...), x_admin_session: Optional[str] = Header(default=None)):
    _require_admin(x_admin_session)

    filename = file.filename or "users.xlsx"
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Upload must be an .xlsx file")

    content = await file.read(USERS_EXCEL_MAX_BYTES + 1)
    if len(content) > USERS_EXCEL_MAX_BYTES:
        raise HTTPException(status_code=413, detail=f"users.xlsx is too large. Maximum size is {USERS_EXCEL_MAX_BYTES} bytes")

    target = Path(USERS_EXCEL_PATH)
    target.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = target.with_name(target.stem + ".upload.tmp.xlsx")

    try:
        # Fail fast if the file is not a valid workbook before touching the live users.xlsx.
        openpyxl.load_workbook(BytesIO(content), read_only=True).close()
        tmp_path.write_bytes(content)
        parsed_count = load_users_from_excel(str(tmp_path), replace_existing=False)
        if parsed_count <= 0:
            raise ValueError("users.xlsx did not contain any valid users")

        tmp_path.replace(target)
        count = load_users_from_excel(str(target), replace_existing=True)
        audit("users_excel_uploaded", detail=f"{filename} uploaded by admin; {count} users loaded")
        return {
            "status": "ok",
            "filename": filename,
            "users_loaded": count,
            "size_bytes": len(content),
            "path": str(target),
        }
    except HTTPException:
        raise
    except Exception as exc:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass
        audit("users_excel_upload_failed", detail=str(exc), status="error")
        raise HTTPException(status_code=400, detail=f"Could not import users.xlsx: {exc}")


@app.post("/admin/reload-users")
async def reload_users(x_admin_session: Optional[str] = Header(default=None)):
    _require_admin(x_admin_session)
    if not os.path.exists(USERS_EXCEL_PATH):
        raise HTTPException(status_code=404, detail=f"Not found: {USERS_EXCEL_PATH}")
    count = load_users_from_excel(USERS_EXCEL_PATH, replace_existing=True)
    audit("users_reloaded", detail=f"{count} users loaded")
    return {"status": "ok", "users_loaded": count}


@app.get("/admin/smtp-test")
async def smtp_test():
    """Sends a test email to the relay account - use to verify Exchange connectivity."""
    html = """<div style="font-family:Arial,sans-serif;padding:24px">
      <p>OTP Relay SMTP test - if you can read this, Exchange is working.</p>
    </div>"""
    try:
        send_email(FROM_EMAIL, "OTP Relay", "OTP Relay - SMTP connectivity test", html)
        return {"status": "ok", "sent_to": FROM_EMAIL}
    except Exception as exc:
        return {"status": "error", "error": str(exc)}


# -- Wizard/admin server-backed endpoints -------------------------------------
@app.get("/admin/auth/status")
async def admin_auth_status():
    return {"configured": bool(_auth_db().get("password_hash"))}


@app.post("/admin/auth/setup")
async def admin_auth_setup(payload: CredentialPayload):
    cred = (payload.credential or "").strip()
    if len(cred) < 4:
        raise HTTPException(status_code=400, detail="Credential too short")
    db = _auth_db()
    if db.get("password_hash"):
        if not payload.current:
            raise HTTPException(status_code=400, detail="Current credential required")
        if not bcrypt.checkpw(payload.current.encode("utf-8"), db["password_hash"].encode("utf-8")):
            raise HTTPException(status_code=401, detail="Current credential incorrect")
    hashed = bcrypt.hashpw(cred.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    _save_auth_db({"password_hash": hashed, "updated_at": _now_iso()})
    session = secrets.token_urlsafe(24)
    ADMIN_SESSIONS[session] = datetime.now(timezone.utc).timestamp()
    audit("admin_auth_setup", detail="Admin credential configured")
    return {"status": "ok", "session": session}


@app.post("/admin/auth/login")
async def admin_auth_login(payload: CredentialPayload):
    db = _auth_db()
    stored = db.get("password_hash")
    if not stored:
        raise HTTPException(status_code=400, detail="Admin credential not configured")
    if not bcrypt.checkpw((payload.credential or "").encode("utf-8"), stored.encode("utf-8")):
        audit("admin_auth_failed", detail="Incorrect admin credential", status="warn")
        raise HTTPException(status_code=401, detail="Incorrect credential")
    session = secrets.token_urlsafe(24)
    ADMIN_SESSIONS[session] = datetime.now(timezone.utc).timestamp()
    audit("admin_auth_login", detail="Admin session opened")
    return {"status": "ok", "session": session}


@app.post("/admin/auth/logout")
async def admin_auth_logout(x_admin_session: Optional[str] = Header(default=None)):
    if x_admin_session:
        ADMIN_SESSIONS.pop(x_admin_session, None)
    return {"status": "ok"}


@app.get("/admin/config")
async def admin_config(x_admin_session: Optional[str] = Header(default=None)):
    _require_admin(x_admin_session)
    return _config_db()


@app.post("/admin/config")
async def admin_config_save(payload: ConfigPayload, x_admin_session: Optional[str] = Header(default=None)):
    _require_admin(x_admin_session)
    tokens: List[str] = []
    for token in payload.admin_tokens:
        clean = str(token or "").strip().upper()
        if clean and clean not in tokens:
            tokens.append(clean)
    _save_config_db({"admin_tokens": tokens, "updated_at": _now_iso()})
    audit("admin_config_saved", detail=f"Configured admin tokens: {', '.join(tokens) or 'none'}")
    return {"status": "ok", "admin_tokens": tokens}


@app.post("/wizard/progress")
async def wizard_progress_save(payload: WizardRecord):
    token = payload.token.strip().upper()
    if token not in users:
        raise HTTPException(status_code=404, detail="Unknown token")
    db = _wizard_db()
    row = _model_dump(payload)
    row["token"] = token
    row["updated_at"] = _now_iso()
    db[token] = row
    _save_wizard_db(db)
    audit("wizard_progress_saved", token=token, detail="Wizard profile/progress updated")
    return {"status": "ok", "record": row}


@app.get("/wizard/progress/{token}")
async def wizard_progress_get(token: str):
    token = token.strip().upper()
    if token not in users:
        raise HTTPException(status_code=404, detail="Unknown token")
    db = _wizard_db()
    return db.get(token, {
        "token": token,
        "display_name": users[token]["name"],
        "iits_username": "",
        "adm_username": "",
        "completed": [],
        "adminCompleted": [],
        "iits_pw_date": None,
        "adm_pw_date": None,
        "vpn_date": None,
    })


@app.get("/admin/wizard")
async def admin_wizard(x_admin_session: Optional[str] = Header(default=None)):
    _require_admin(x_admin_session)
    db = _wizard_db()
    merged = []
    for token, user in sorted(users.items()):
        rec = db.get(token, {})
        merged.append({
            "token": token,
            "display_name": rec.get("display_name") or user.get("name", ""),
            "email": user.get("email", ""),
            "iits_username": rec.get("iits_username", ""),
            "adm_username": rec.get("adm_username", ""),
            "completed": rec.get("completed", []),
            "adminCompleted": rec.get("adminCompleted", []),
            "iits_pw_date": rec.get("iits_pw_date"),
            "adm_pw_date": rec.get("adm_pw_date"),
            "vpn_date": rec.get("vpn_date"),
            "updated_at": rec.get("updated_at"),
        })
    return {"users": merged}


@app.post("/api/onboard/notify")
async def onboard_notify(request: Request):
    payload = await request.json()
    token = str(payload.get("token", "") or "").strip().upper() or None
    detail = json.dumps(payload, sort_keys=True)[:500]
    audit("onboard_notify", token=token, detail=detail)
    return {"status": "ok", "received": payload, "ts": _now_iso()}


@app.get("/guide.html", include_in_schema=False)
def serve_guide_html():
    guide_path = FRONTEND_DIR / "guide.html"
    if not guide_path.exists():
        raise HTTPException(status_code=404, detail="guide.html not deployed")
    return FileResponse(guide_path)


# Serve frontend - must be last
app.mount("/", StaticFiles(directory=str(FRONTEND_DIR), html=True), name="frontend")
